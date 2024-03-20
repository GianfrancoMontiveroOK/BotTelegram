import openpyxl

def agregar_datos_a_excel(user,cardNum,id):
    archivo="datos.xlsx"

    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
        print("access to DB")
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
        print("Create DB")
    
    # Selecciona la primera hoja del libro de trabajo o crea una nueva
    sheet = workbook.active

   ###################################################################################
    # Itera sobre cada fila en la hoja
   
    fila_vacia = sheet.max_row + 1
    
    # Guarda los datos en la primera fila vacía
    sheet.cell(row=fila_vacia, column=1).value = id
    sheet.cell(row=fila_vacia, column=2).value = user
    sheet.cell(row=fila_vacia, column=3).value = cardNum
    sheet.cell(row=fila_vacia, column=4).value = 0

    
    # Guarda los cambios en el archivo
    workbook.save(archivo)
    print(f"El número '{id}' Junto con su USER y su WALLET Se agregó en la fila {fila_vacia}.")
def auth(id):
    archivo="datos.xlsx"
    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook.active
        print("access DB /AUTH")
        # Itera sobre cada fila en la hoja
        for fila in sheet.iter_rows(values_only=True):
            if fila[0] == id:
                print("USER /AUTH")
                return True
        print("DB /AUTH NOT FOUND")
        return False      
    except FileNotFoundError:
        print("error access DB /AUTH")
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
    # Selecciona la primera hoja del libro de trabajo
    ###################################################################################
def reffPlus(id):
    print("Referrer ID:", id)
    archivo="datos.xlsx"
    id_def=id
    
    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
        # Selecciona la primera hoja del libro de trabajo o crea una nueva
        sheet = workbook.active
        # Itera sobre cada fila en la hoja
        
        for fila_num, fila in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Verifica si fila[0] es un número válido
            if isinstance(fila[0], int) or fila[0].isdigit():
                value = int(fila[0])
                value2 = int(id_def)
                
                if value == value2:
                    cell_value = sheet.cell(row=fila_num, column=4).value
                    # Verifica si el valor de la celda no es None
                    if cell_value is not None:
                        # Si el valor de la celda no es None, suma 1 al valor existente
                        sheet.cell(row=fila_num, column=4).value = cell_value + 1
                        print("Set Refferal into DB")
                        workbook.save(archivo)
                    else:
                        # Si el valor de la celda es None, asigna 1 a la celda
                        sheet.cell(row=fila_num, column=4).value = 1
                        print("Set Refferal into DB")
                        workbook.save(archivo)                                    
                else:
                    print("iter..")
            else:
                print("iter..")
                        
        return
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
        print("error DB /REFF++")
        return
def reffCheck(id):
    archivo="datos.xlsx"
    id_def=id
    
    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
    except:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
    # Selecciona la primera hoja del libro de trabajo o crea una nueva
    sheet = workbook.active
   ###################################################################################
    for fila_num, fila in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Verifica si fila[0] es un número válido
            if isinstance(fila[0], int) or fila[0].isdigit():
                value = int(fila[0])
                value2 = int(id_def)
                
                if value == value2:
                    cell_value = sheet.cell(row=fila_num, column=4).value
                    # Verifica si el valor de la celda no es None
                    if cell_value is not None:
                        # Si el valor de la celda no es None, suma 1 al valor existente
                        reffNum = sheet.cell(row=fila_num, column=4).value = cell_value + 1
                        return reffNum-1
                    else:
                        return "0"                                  
def ironCheck(id):
    archivo="datos.xlsx"
    id_def=id
    
    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
    except:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
    # Selecciona la primera hoja del libro de trabajo o crea una nueva
    sheet = workbook.active
   ###################################################################################
    for fila_num, fila in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Verifica si fila[0] es un número válido
            if isinstance(fila[0], int) or fila[0].isdigit():
                value = int(fila[0])
                value2 = int(id_def)
                
                if value == value2:
                    cell_value = sheet.cell(row=fila_num, column=4).value
                    # Verifica si el valor de la celda no es None
                    if cell_value is not None:
                        # Si el valor de la celda no es None, calcular cantidad iron
                        data = 10 + cell_value * 5
                        return data
                    else:
                        return "0"                                  
def verifyCheck(id):
    archivo="datos.xlsx"
    id_def=id
    
    try:
        # Intenta cargar el archivo existente
        workbook = openpyxl.load_workbook(archivo)
    except:
        # Si el archivo no existe, crea uno nuevo
        workbook = openpyxl.Workbook()
    # Selecciona la primera hoja del libro de trabajo o crea una nueva
    sheet = workbook.active
   ###################################################################################
    for fila_num, fila in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Verifica si fila[0] es un número válido
            if isinstance(fila[0], int) or fila[0].isdigit():
                value = int(fila[0])
                value2 = int(id_def)
                
                if value == value2:
                    wallet_value = sheet.cell(row=fila_num, column=3).value
                    user_value = sheet.cell(row=fila_num, column=2).value
                    # Verifica si el valor de la celda no es None
                    if user_value is not None and wallet_value is not None:
                        sheet.cell(row=fila_num, column=5).value = "VERIFY"
                        return (wallet_value,user_value)
                    else:
                        return "0"                                  
                else:
                    print("RFFCHECK ELSE", fila[0])
            else:
                print("El valor de la primera columna en la fila {} no es un número válido: {}".format(fila_num, fila[0]))